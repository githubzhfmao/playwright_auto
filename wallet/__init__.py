from openpyxl import *


class Excel:

    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    def get_columns(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    def get_rows(self):
        return self.ws.max_row


def get_wallet(file):
    # 用户助记词路径，以xlsx格式保存，该路径由用户提供,在config.ini中配置
    # file = global_config.get('path', 'wallet_path').strip()
    address_list = Excel(file).get_columns(1)
    mnemonic_list = Excel(file).get_columns(3)
    wallet = [address_list, mnemonic_list]
    return wallet


def get_mnemonic(file):
    # 用户助记词路径，以xlsx格式保存，该路径由用户提供,在config.ini中配置
    # file = global_config.get('path', 'wallet_path').strip()
    mnemonic_list = Excel(file).get_columns(1)
    return mnemonic_list
